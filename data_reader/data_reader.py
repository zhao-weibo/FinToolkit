import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
from openpyxl import load_workbook


class FileAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Analyzer (Excel/CSV/DTA)")

        # Create select file button
        self.select_file_btn = tk.Button(root, text="Select File (Excel/CSV/DTA)", command=self.load_file)
        self.select_file_btn.pack(pady=10)

        # Display file information
        self.info_text = tk.Text(root, width=80, height=5, wrap='word', state='disabled')
        self.info_text.pack(padx=10, pady=10)

        # Create show first 1000 rows button
        self.show_first_btn = tk.Button(root, text="Show First 1000 Rows", command=self.show_first_preview, state='disabled')
        self.show_first_btn.pack(pady=5)

        # Create show last 1000 rows button
        self.show_last_btn = tk.Button(root, text="Show Last 1000 Rows", command=self.show_last_preview, state='disabled')
        self.show_last_btn.pack(pady=5)

        self.file_path = None  # Store full file path for reference
        self.file_type = None  # Store the file type for logic handling
        self.total_rows = None  # Store total rows (if calculated)

    def load_file(self):
        # Reset state
        self.info_text.config(state='normal')
        self.info_text.delete(1.0, tk.END)
        self.info_text.config(state='disabled')
        self.show_first_btn.config(state='disabled')
        self.show_last_btn.config(state='disabled')
        self.file_path = None
        self.file_type = None

        # Open file dialog
        file_path = filedialog.askopenfilename(
            title="Select File (Excel/CSV/DTA)",
            filetypes=[
                ("CSV Files", "*.csv"),
                ("Excel Files", "*.xls;*.xlsx"),
                ("Stata Files", "*.dta")
            ]
        )
        if not file_path:
            return

        self.file_path = file_path
        if file_path.endswith('.csv'):
            self.file_type = 'csv'
        elif file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            self.file_type = 'excel'
        elif file_path.endswith('.dta'):
            self.file_type = 'stata'
        else:
            messagebox.showerror("Error", "Unsupported file type selected.")
            return

        try:
            # Display file information
            self.display_info()
            self.show_first_btn.config(state='normal')
            self.show_last_btn.config(state='normal')
        except Exception as e:
            messagebox.showerror("Error", f"Error processing file: {e}")

    def display_info(self):
        """Display basic information about the file."""
        self.info_text.config(state='normal')
        self.info_text.delete(1.0, tk.END)

        # File-specific information
        info = f"File Path: {self.file_path}\n"
        info += f"File Type: {self.file_type.capitalize()}\n"
        self.info_text.insert(tk.END, info)

        self.info_text.config(state='disabled')

    def show_first_preview(self):
        """Show the first 1000 rows of the file."""
        try:
            if self.file_type == 'csv':
                preview = pd.read_csv(self.file_path, nrows=1000)
            elif self.file_type == 'excel':
                preview = pd.read_excel(self.file_path, nrows=1000)
            elif self.file_type == 'stata':
                # Use chunksize to read the first 1000 rows
                preview = next(pd.read_stata(self.file_path, chunksize=1000))
            else:
                raise ValueError("Unsupported file type.")
            self.show_preview_window(preview, "First 1000 Rows of Data")
        except Exception as e:
            messagebox.showerror("Error", f"Error loading first 1000 rows: {e}")

    def show_last_preview(self):
        """Show the last 1000 rows of the file."""
        try:
            if self.file_type == 'csv':
                preview = self.get_last_n_rows_csv(self.file_path, n=1000)
            elif self.file_type == 'excel':
                preview = self.get_last_n_rows_excel(self.file_path, n=1000)
            elif self.file_type == 'stata':
                preview = self.get_last_n_rows_stata(self.file_path, n=1000)
            else:
                raise ValueError("Unsupported file type.")
            self.show_preview_window(preview, "Last 1000 Rows of Data")
        except Exception as e:
            messagebox.showerror("Error", f"Error loading last 1000 rows: {e}")

    def get_last_n_rows_csv(self, file_path, n=1000):
        """Efficiently get the last N rows of a CSV file."""
        with open(file_path, 'rb') as f:
            f.seek(0, 2)  # Move to the end of the file
            buffer = bytearray()
            pointer_location = f.tell()
            while pointer_location > 0 and len(buffer) < n * 10000:  # Approximate size
                pointer_location -= 1024
                f.seek(max(pointer_location, 0))
                buffer = f.read(1024) + buffer
            lines = buffer.decode(errors='ignore').splitlines()
            return pd.DataFrame([line.split(',') for line in lines[-n:]])

    def get_last_n_rows_excel(self, file_path, n=1000):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        total_rows = wb.active.max_row
        skip_rows = max(0, total_rows - n)
        return pd.read_excel(file_path, skiprows=skip_rows)

    def get_last_n_rows_stata(self, file_path, n=1000):
        """Get the last N rows of a Stata file."""
        try:
            # Read file in chunks
            chunks = pd.read_stata(file_path, chunksize=n)
            last_chunk = None
            for chunk in chunks:
                last_chunk = chunk  # Keep updating to the last chunk
            return last_chunk
        except Exception as e:
            raise ValueError(f"Error while reading the last {n} rows from Stata file: {e}")

    def show_preview_window(self, preview, title):
        """Show the given preview in a new window."""
        preview_window = tk.Toplevel(self.root)
        preview_window.title(title)

        # Set fixed window size (e.g., width=800, height=600)
        window_width = 800
        window_height = 600

        # Center the window on the screen
        screen_width = preview_window.winfo_screenwidth()
        screen_height = preview_window.winfo_screenheight()
        x_cord = int((screen_width / 2) - (window_width / 2))
        y_cord = int((screen_height / 2) - (window_height / 2))
        preview_window.geometry(f"{window_width}x{window_height}+{x_cord}+{y_cord}")

        # Create a frame to hold the Treeview and Scrollbars
        frame = tk.Frame(preview_window)
        frame.pack(fill=tk.BOTH, expand=True)

        # Create the Treeview
        tree = ttk.Treeview(frame, show="headings")
        tree.grid(row=0, column=0, sticky="nsew")  # Use grid to manage layout

        # Add vertical scrollbar
        y_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        y_scrollbar.grid(row=0, column=1, sticky="ns")  # Attach to the right side of the Treeview
        tree.configure(yscrollcommand=y_scrollbar.set)

        # Add horizontal scrollbar
        x_scrollbar = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        x_scrollbar.grid(row=1, column=0, sticky="ew")  # Attach to the bottom of the Treeview
        tree.configure(xscrollcommand=x_scrollbar.set)

        # Ensure the frame expands with the window
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        # Set column headers
        tree["columns"] = list(preview.columns)
        for col in preview.columns:
            tree.heading(col, text=col)
            tree.column(col, anchor='center', width=150)  # Adjust column width as needed

        # Insert data
        for _, row in preview.iterrows():
            tree.insert("", "end", values=list(row))


if __name__ == "__main__":
    root = tk.Tk()
    app = FileAnalyzerApp(root)
    root.mainloop()
